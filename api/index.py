from flask import Flask, render_template, request, send_file, redirect, url_for
import openpyxl

app = Flask(__name__)

def get_excel_file_path():
    return '/tmp/students.xlsx'

def get_workbook(excel_file_path):
    try:
        return openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        return openpyxl.Workbook()

def get_worksheet(workbook):
    return workbook.active

def get_headers():
    return ['Name', 'Age', 'Grade']

def append_headers(worksheet):
    worksheet.append(get_headers())

def append_data(worksheet, data):
    worksheet.append(data)

def save_workbook(excel_file_path, workbook):
    workbook.save(excel_file_path)

def save_student_to_excel(student_data, excel_file_path):
    workbook = get_workbook(excel_file_path)
    worksheet = get_worksheet(workbook)

    if not worksheet.dimensions:
        append_headers(worksheet)

    data = []
    for field in get_headers():
        data.append(student_data.get(field, ''))
    
    append_data(worksheet, data)
    save_workbook(excel_file_path, workbook)

def read_students_from_excel(excel_file_path):
    workbook = get_workbook(excel_file_path)
    worksheet = get_worksheet(workbook)

    if not worksheet.dimensions:
        return []

    headers = get_headers()
    students_data = []
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        students_data.append(dict(zip(headers, row)))

    return students_data

@app.route('/')
def index():
    excel_file_path = get_excel_file_path()
    return render_elements(excel_file_path)

@app.route('/static/<path:path>')
def static_file(path):
    return app.send_static_file(path)

@app.route('/add_student', methods=['POST'])
def add_student():
    if request.method == 'POST':
        student_data = request.form.to_dict()
        excel_file_path = get_excel_file_path()
        save_student_to_excel(student_data, excel_file_path)
        # Redirect to the home page after processing the form
        return redirect(url_for('index'))
    return render_elements(get_excel_file_path())

def render_elements(excel_file_path):
    return render_template('index.html', students_data=read_students_from_excel(excel_file_path))

@app.route('/clear_data')
def clear_data():
    excel_file_path = get_excel_file_path()
    workbook = openpyxl.Workbook()
    save_workbook(excel_file_path, workbook)
    return render_elements(excel_file_path)

@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    if request.method == 'POST':
        input_data = request.form.to_dict()
        excel_file_path = get_excel_file_path()
        save_student_to_excel(input_data, excel_file_path)
        return send_file(excel_file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
